import json
import pandas as pd
from collections import Counter
from pathlib import Path
from openpyxl.styles import PatternFill

def is_alphabet(gloss):
    """Check if gloss is a single letter (A-Z or a-z)"""
    return len(gloss) == 1 and gloss.isalpha()

def generate_statistics():
    """
    Extract gloss, source, and video_id from exists.json and generate an Excel file
    with statistics.
    """
    
    # Load the JSON file
    json_file = Path('.\\exists.json')
    print(f"json file path: {json_file}")
    
    if not json_file.exists():
        print(f"Error: {json_file} not found")
        return
    
    with open(json_file, 'r') as f:
        data = json.load(f)
    
    # Extract data: gloss, source, video_id
    records = []
    
    for gloss_entry in data:
        gloss = gloss_entry.get('gloss', 'Unknown')
        instances = gloss_entry.get('instances', [])
        
        for instance in instances:
            record = {
                'Gloss': gloss,
                'Source': instance.get('source', 'Unknown'),
                'Video_ID': instance.get('video_id', 'Unknown')
            }
            records.append(record)
    
    # Create DataFrame
    df = pd.DataFrame(records)
    
    # Generate statistics
    stats_data = {
        'Statistic': [],
        'Count': []
    }
    
    # Total entries
    stats_data['Statistic'].append('Total Entries')
    stats_data['Count'].append(len(df))
    
    # Unique glosses
    stats_data['Statistic'].append('Unique Glosses')
    stats_data['Count'].append(df['Gloss'].nunique())
    
    # Unique sources
    stats_data['Statistic'].append('Unique Sources')
    stats_data['Count'].append(df['Source'].nunique())
    
    # Unique video IDs
    stats_data['Statistic'].append('Unique Video IDs')
    stats_data['Count'].append(df['Video_ID'].nunique())
    
    stats_df = pd.DataFrame(stats_data)
    
    # Create source distribution
    source_counts = df['Source'].value_counts().reset_index()
    source_counts.columns = ['Source', 'Count']
    
    # Create gloss distribution
    gloss_counts = df['Gloss'].value_counts().reset_index()
    gloss_counts.columns = ['Gloss', 'Count']
    
    # Create gloss-source matrix (each source as a column)
    gloss_source_matrix = pd.crosstab(df['Gloss'], df['Source'], margins=False)
    gloss_source_matrix = gloss_source_matrix.reset_index()
    gloss_source_matrix.columns.name = None
    
    # Create gloss-source matrix (each source as a column)
    gloss_source_matrix = pd.crosstab(df['Gloss'], df['Source'], margins=False)
    gloss_source_matrix = gloss_source_matrix.reset_index()
    gloss_source_matrix.columns.name = None
    
    # Write to Excel with multiple sheets
    output_file = 'statistics.xlsx'
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet 1: Raw data
        df.to_excel(writer, sheet_name='Raw Data', index=False)
        
        # Sheet 2: Statistics summary
        stats_df.to_excel(writer, sheet_name='Summary Statistics', index=False)
        
        # Sheet 3: Source distribution
        source_counts.to_excel(writer, sheet_name='Source Distribution', index=False)
        
        # Sheet 4: Gloss distribution
        gloss_counts.to_excel(writer, sheet_name='Gloss Distribution', index=False)
        
        # Sheet 5: Gloss-Source Matrix
        gloss_source_matrix.to_excel(writer, sheet_name='Gloss-Source Matrix', index=False)
        
        # Apply red formatting for alphabets in Gloss Distribution sheet
        gloss_dist_worksheet = writer.sheets['Gloss Distribution']
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        for row_idx, row in enumerate(gloss_dist_worksheet.iter_rows(min_row=2, max_row=gloss_dist_worksheet.max_row), start=2):
            gloss_value = row[0].value
            if gloss_value and is_alphabet(str(gloss_value)):
                row[0].fill = red_fill
        
        # Apply red formatting for alphabets in Gloss-Source Matrix sheet
        gloss_source_worksheet = writer.sheets['Gloss-Source Matrix']
        
        for row_idx, row in enumerate(gloss_source_worksheet.iter_rows(min_row=2, max_row=gloss_source_worksheet.max_row), start=2):
            gloss_value = row[0].value
            if gloss_value and is_alphabet(str(gloss_value)):
                row[0].fill = red_fill
    
    print(f"✓ Excel file '{output_file}' created successfully!")
    print(f"\nStatistics Summary:")
    print(f"  Total Entries: {len(df)}")
    print(f"  Unique Glosses: {df['Gloss'].nunique()}")
    print(f"  Unique Sources: {df['Source'].nunique()}")
    print(f"  Unique Video IDs: {df['Video_ID'].nunique()}")


if __name__ == '__main__':
    generate_statistics()