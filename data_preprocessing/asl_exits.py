import json
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- Paths Configuration ---
JSON_INPUT_PATH = "asl_words.json"
MS_VIDEO_DIR = "Microsoft_Videos"
OTHER_VIDEO_DIR = "videos"

JSON_FOUND_PATH = "asl_mis_wlasl.json"
JSON_MISSING_PATH = "missing_v2.json"
EXCEL_OUTPUT_PATH = "asl_mis_wlasl.xlsx"


def build_video_cache(directory, is_microsoft=False):
    """Reads all filenames under the target folder.

    If it's Microsoft, it extracts only the leading digits from filenames like
    '12345-gloss.mp4'.
    """
    if not os.path.exists(directory):
        return set()

    cache = set()
    for f in os.listdir(directory):
        base_name = os.path.splitext(f)[0]  # remove .mp4 extension
        if is_microsoft:

            # using re.match to find the leading digits in the filename
            match = re.match(r"^(\d+)", base_name)
            if match:
                cache.add(match.group(1))  # only add the leading digits to the cache
        else:
            cache.add(base_name)  # other databases keep the original names

    return cache


def main():
    # 1. Load input JSON file
    with open(JSON_INPUT_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Performance Optimization: Cache video filenames in memory
    print("Building video file cache...")
    ms_video_cache = build_video_cache(MS_VIDEO_DIR, is_microsoft=True)
    other_video_cache = build_video_cache(OTHER_VIDEO_DIR, is_microsoft=False)
    print("Cache built successfully. Starting data comparison...")

    combined_records = []  # merge found and missing records into one list with status field

    # 2. Loop through data and check if the video exists
    for item in data:
        gloss = item.get("gloss")
        video_id = str(item.get("video_id"))  # Ensure string type
        source = item.get("source")

        # Determine which memory cache to use based on the source
        if str(source).lower() == "microsoft":
            is_found = video_id in ms_video_cache
        else:
            is_found = video_id in other_video_cache

        # add status field to indicate whether the video was found or missing
        status_str = "Found" if is_found else "Missing"
        
        record = {
            "gloss": gloss, 
            "video_id": video_id, 
            "source": source,
            "status": status_str  
        }
        
        combined_records.append(record)

    # save the combined records to a new JSON file for reference
    with open(JSON_FOUND_PATH, "w", encoding="utf-8") as f:
        json.dump(combined_records, f, indent=4, ensure_ascii=False)

    print(f"JSON processing completed. Total combined records: {len(combined_records)}")

    # 6. Generate Excel file and statistics
    df_combined = pd.DataFrame(combined_records)

    if df_combined.empty:
        print("No data to process. Excel generation aborted.")
        return

    # Stat 1: Classify source types (microsoft vs not_microsoft)
    df_combined["source_type"] = df_combined["source"].apply(
        lambda x: "microsoft" if str(x).lower() == "microsoft" else "not_microsoft"
    )

    # Create pivot table (contains all the words, even those with zero counts in either category)
    stat_source = (
        df_combined.groupby(["gloss", "source_type"])
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )

    # Ensure both columns exist in the DataFrame
    if "microsoft" not in stat_source.columns:
        stat_source["microsoft"] = 0
    if "not_microsoft" not in stat_source.columns:
        stat_source["not_microsoft"] = 0

    # Reorder and rename columns
    stat_source = stat_source[["gloss", "microsoft", "not_microsoft"]]
    stat_source.columns = ["Gloss", "Microsoft Count", "Not Microsoft Count"]

    # Stat 2: Count total unique glosses
    total_unique_gloss = df_combined["gloss"].nunique()
    
    # extra calculation: count total found and missing videos
    total_found = (df_combined["status"] == "Found").sum()
    total_missing = (df_combined["status"] == "Missing").sum()

    # Create summary DataFrame
    df_summary = pd.DataFrame(
        {
            "Metrics": [
                "Total Unique Gloss Count", 
                "Total Videos Found", 
                "Total Videos Missing"
            ],
            "Values": [total_unique_gloss, total_found, total_missing],
        }
    )

    # Write data frames to separate Excel sheets
    with pd.ExcelWriter(EXCEL_OUTPUT_PATH, engine="openpyxl") as writer:
        # Details sheet includes all records with their status (Found/Missing)
        df_combined[["gloss", "video_id", "source", "status"]].to_excel(
            writer, sheet_name="Details", index=False
        )
        stat_source.to_excel(
            writer, sheet_name="Gloss Source Stats", index=False
        )
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

    # Stat 3: Highlight Gloss columns based on conditions
    wb = load_workbook(EXCEL_OUTPUT_PATH)

    light_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    light_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    num_words = {"one", "two", "three", "four", "five", "six", "seven", "eight", "nine", "ten", "zero"}

    for sheet_name in ["Details", "Gloss Source Stats"]:
        ws = wb[sheet_name]
        gloss_col_idx = None
        for cell in ws[1]:
            if cell.value and str(cell.value).lower() == "gloss":
                gloss_col_idx = cell.column
                break

        if gloss_col_idx is None:
            continue

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=gloss_col_idx)
            val = str(cell.value).strip() if cell.value is not None else ""

            tokens = set(re.split(r"[^a-zA-Z0-9]", val.lower()))
            has_digit = any(char.isdigit() for char in val)
            has_num_word = not tokens.isdisjoint(num_words)
            is_single_letter = len(val) == 1 and val.isalpha()

            if has_digit or has_num_word:
                cell.fill = light_blue
            elif is_single_letter:
                cell.fill = light_red

    wb.save(EXCEL_OUTPUT_PATH)
    print(f"Excel file generated successfully: {EXCEL_OUTPUT_PATH}")


if __name__ == "__main__":
    main()