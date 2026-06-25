import json
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

'''
This module is used for checking if the video_ids exist under the Microsoft_Videos and videos folder.
By reading the microsoft.json file and exists_v2.json file, we will check if the video_id exists under the corresponding folders, and add a status field to indicate whether the video was found or missing.
Then we will save the result as asl_mis_wlasl.json, which contains all the records with their status. 
We will also generate an Excel file asl_mis_wlasl.xlsx to summarize the statistics of the found and missing videos, and highlight the glosses based on certain conditions 
(e.g., if the gloss contains digits or number words, or if it's a single letter).
'''


# merge microsoft.json and exists_v2.json into a single list of records
# if a gloss exists in both files, merge the items
# if a gloss exists in only one file, add it to the merged list
def merge_jsons(mis_path, wlasl_path):
    print(f"Merging {mis_path} and {wlasl_path}...")
    '''
    Merge the mis_data and wlasl_data into a single list of records.
    '''
    # 1. Load input JSON file
    with open(mis_path, "r", encoding="utf-8") as f:
        mis_data = json.load(f)
    with open(wlasl_path, "r", encoding="utf-8") as f:
        wlasl_data = json.load(f)

    merged_by_gloss = {}

    def format_item(item):
        return {
            'video_id': item.get('video_id', ''),
            'source': item.get('source', '')
        }

    # Process Microsoft data
    for mis in mis_data:
        gloss_upper = mis.get('gloss', '').upper()
        formatted_items = [format_item(i) for i in mis.get('item', [])]
        
        if gloss_upper in merged_by_gloss:
            merged_by_gloss[gloss_upper]['item'].extend(formatted_items)
        else:
            merged_by_gloss[gloss_upper] = {
                'gloss': gloss_upper,
                'item': formatted_items
            }

    # Process WLASL data
    for wlasl in wlasl_data:
        gloss_upper = wlasl.get('gloss', '').upper()
        formatted_items = [format_item(i) for i in wlasl.get('item', [])]
        
        if gloss_upper in merged_by_gloss:
            merged_by_gloss[gloss_upper]['item'].extend(formatted_items)
        else:
            merged_by_gloss[gloss_upper] = {
                'gloss': gloss_upper,
                'item': formatted_items
            }

    return list(merged_by_gloss.values())


def build_video_cache(directory):
    """
    Reads all filenames under the target folder.
    """
    if not os.path.exists(directory):
        return set()

    cache = set()
    for f in os.listdir(directory):
        base_name = os.path.splitext(f)[0]  # remove .mp4 extension
        cache.add(base_name)

    return cache


def main():

    # Performance Optimization: Cache video filenames in memory
    print("Building video file cache...")
    ms_video_cache = build_video_cache(MS_VIDEO_DIR)
    other_video_cache = build_video_cache(OTHER_VIDEO_DIR)
    print("Cache built successfully. Starting data comparison...")

    combined_records = []  # merge found and missing records into one list with status field

    print("Merging JSON files...")
    merged_data = merge_jsons(MIS_JSON_PATH, WLASL_JSON_PATH)

    # 2. Loop through data and check if the video exists
    for entry in merged_data:
        gloss = entry.get("gloss", "")
        items = entry.get("item", [])
        combined_records.append({
            'gloss': gloss,
            "status":"",
            'item': []
        })

        for item in items:
            video_id = item.get("video_id", "")
            source = item.get("source", "")
            
            if source == 'microsoft':
                is_found = os.path.splitext(video_id)[0] in ms_video_cache
            else:
                is_found=video_id in other_video_cache
            
            # Add status to the merged item
            item["status"] = "Found" if is_found else "Missing"
            
            # Flatten to combined_records for Excel generation
            combined_records.append({
                'gloss': gloss,
                'video_id': video_id,
                'source': source,
                'status': item["status"]
            })

    # save the combined records to a new JSON file for reference
    with open(JSON_FOUND_PATH, "w", encoding="utf-8") as f:
        json.dump(merged_data, f, indent=4, ensure_ascii=False)

    print(f"JSON processing completed. Total combined records: {len(combined_records)}")

    # 6. Generate Excel file and statistics
    df_combined = pd.DataFrame(combined_records)

    if df_combined.empty:
        print("No data to process. Excel generation aborted.")
        return

    # Stat 1: Classify source types (microsoft vs not_microsoft)
    df_combined["source_type"] = df_combined["source"].apply(
        lambda x: "microsoft" if str(x).lower() == "microsoft" else "not_microsoft"
    )

    # Create pivot table (contains all the words, even those with zero counts in either category)
    stat_source = (
        df_combined.groupby(["gloss", "source_type"])
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )

    # Ensure both columns exist in the DataFrame
    if "microsoft" not in stat_source.columns:
        stat_source["microsoft"] = 0
    if "not_microsoft" not in stat_source.columns:
        stat_source["not_microsoft"] = 0

    # Reorder and rename columns
    stat_source = stat_source[["gloss", "microsoft", "not_microsoft"]]
    stat_source.columns = ["Gloss", "Microsoft Count", "Not Microsoft Count"]

    # Stat 2: Count total unique glosses
    total_unique_gloss = df_combined["gloss"].nunique()
    
    # extra calculation: count total found and missing videos
    total_found = (df_combined["status"] == "Found").sum()
    total_missing = (df_combined["status"] == "Missing").sum()

    # Create summary DataFrame
    df_summary = pd.DataFrame(
        {
            "Metrics": [
                "Total Unique Gloss Count", 
                "Total Videos Found", 
                "Total Videos Missing"
            ],
            "Values": [total_unique_gloss, total_found, total_missing],
        }
    )

    # Write data frames to separate Excel sheets
    with pd.ExcelWriter(EXCEL_OUTPUT_PATH, engine="openpyxl") as writer:
        # Details sheet includes all records with their status (Found/Missing)
        df_combined[["gloss", "video_id", "source", "status"]].to_excel(
            writer, sheet_name="Details", index=False
        )
        stat_source.to_excel(
            writer, sheet_name="Gloss Source Stats", index=False
        )
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

    # Stat 3: Highlight Gloss columns based on conditions
    wb = load_workbook(EXCEL_OUTPUT_PATH)

    light_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    light_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    num_words = {"one", "two", "three", "four", "five", "six", "seven", "eight", "nine", "ten", "zero"}

    for sheet_name in ["Details", "Gloss Source Stats"]:
        ws = wb[sheet_name]
        gloss_col_idx = None
        for cell in ws[1]:
            if cell.value and str(cell.value).lower() == "gloss":
                gloss_col_idx = cell.column
                break

        if gloss_col_idx is None:
            continue

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=gloss_col_idx)
            val = str(cell.value).strip() if cell.value is not None else ""

            tokens = set(re.split(r"[^a-zA-Z0-9]", val.lower()))
            has_digit = any(char.isdigit() for char in val)
            has_num_word = not tokens.isdisjoint(num_words)
            is_single_letter = len(val) == 1 and val.isalpha()

            if has_digit or has_num_word:
                cell.fill = light_blue
            elif is_single_letter:
                cell.fill = light_red

    wb.save(EXCEL_OUTPUT_PATH)
    print(f"Excel file generated successfully: {EXCEL_OUTPUT_PATH}")


if __name__ == "__main__":
    
    # --- Paths Configuration ---
    MIS_JSON_PATH = "..\\data_preprocessing\\microsoft.json"
    WLASL_JSON_PATH = "..\\data_preprocessing\\exists_v2.json"

    MS_VIDEO_DIR = "Microsoft_Videos"
    OTHER_VIDEO_DIR = "videos"

    JSON_FOUND_PATH = "..\\data_preprocessing\\asl_mis_wlasl.json"
    EXCEL_OUTPUT_PATH = "..\\data_preprocessing\\asl_mis_wlasl.xlsx"
    # main()
    merged_json= merge_jsons(MIS_JSON_PATH, WLASL_JSON_PATH)

    print(json.dumps(merged_json, indent=4, ensure_ascii=False))