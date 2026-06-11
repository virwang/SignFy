import json
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- Paths Configuration ---
JSON_INPUT_PATH = "asl_words.json"
MS_VIDEO_DIR = "Microsoft_Videos"
OTHER_VIDEO_DIR = "videos"

JSON_FOUND_PATH = "asl_mis_wlasl.json"
JSON_MISSING_PATH = "missing_v2.json"
EXCEL_OUTPUT_PATH = "asl_mis_wlasl.xlsx"


def build_video_cache(directory):
    """Reads all filenames under the target folder and extracts the base

    filenames (without extensions) into a set to enable O(1) lookups.
    """
    if not os.path.exists(directory):
        return set()

    return {os.path.splitext(f)[0] for f in os.listdir(directory)}


def main():
    # 1. Load input JSON file
    with open(JSON_INPUT_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Performance Optimization: Cache video filenames in memory before the loop
    print("Building video file cache...")
    ms_video_cache = build_video_cache(MS_VIDEO_DIR)
    other_video_cache = build_video_cache(OTHER_VIDEO_DIR)
    print("Cache built successfully. Starting data comparison...")

    found_records = []
    missing_records = []

    # 2. & 5. Loop through data and check if the video exists
    for item in data:
        gloss = item.get("gloss")
        video_id = str(item.get("video_id"))  # Ensure string type for comparison
        source = item.get("source")

        # Determine which memory cache to use based on the source
        if str(source).lower() == "microsoft":
            is_found = video_id in ms_video_cache
        else:
            is_found = video_id in other_video_cache

        record = {"gloss": gloss, "video_id": video_id, "source": source}

        if is_found:
            # Found in cache: keep the first match by appending directly
            found_records.append(record)
        else:
            # Missing record
            missing_records.append(record)

    # 3. Save matching records to JSON
    with open(JSON_FOUND_PATH, "w", encoding="utf-8") as f:
        json.dump(found_records, f, indent=4, ensure_ascii=False)

    # 4. Save missing records to JSON
    with open(JSON_MISSING_PATH, "w", encoding="utf-8") as f:
        json.dump(missing_records, f, indent=4, ensure_ascii=False)

    print(
        f"JSON processing completed. Found: {len(found_records)}, Missing: {len(missing_records)}"
    )

    # 6. Generate Excel file and statistics
    df_found = pd.DataFrame(found_records)

    if df_found.empty:
        print("No videos found. Excel generation aborted.")
        return

    # Stat 1: Classify source types (microsoft vs not_microsoft)
    df_found["source_type"] = df_found["source"].apply(
        lambda x: "microsoft" if str(x).lower() == "microsoft" else "not_microsoft"
    )

    # Create pivot table to count occurrences per gloss per source type
    stat_source = (
        df_found.groupby(["gloss", "source_type"])
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )

    # Ensure both columns exist in the DataFrame
    if "microsoft" not in stat_source.columns:
        stat_source["microsoft"] = 0
    if "not_microsoft" not in stat_source.columns:
        stat_source["not_microsoft"] = 0

    # Reorder and rename columns
    stat_source = stat_source[["gloss", "microsoft", "not_microsoft"]]
    stat_source.columns = ["Gloss", "Microsoft Count", "Not Microsoft Count"]

    # Stat 2: Count total unique glosses
    total_unique_gloss = df_found["gloss"].nunique()

    # Create summary DataFrame
    df_summary = pd.DataFrame(
        {
            "Metrics": ["Total Unique Gloss Count"],
            "Values": [total_unique_gloss],
        }
    )

    # Write data frames to separate Excel sheets
    with pd.ExcelWriter(EXCEL_OUTPUT_PATH, engine="openpyxl") as writer:
        df_found[["gloss", "video_id", "source"]].to_excel(
            writer, sheet_name="Details", index=False
        )
        stat_source.to_excel(
            writer, sheet_name="Gloss Source Stats", index=False
        )
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

    # Stat 3: Highlight Gloss columns based on conditions
    wb = load_workbook(EXCEL_OUTPUT_PATH)

    # Define fills
    light_blue = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )  # Light Blue for numbers
    light_red = PatternFill(
        start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
    )  # Light Red for single letters

    # Number words dictionary for token matching
    num_words = {
        "one",
        "two",
        "three",
        "four",
        "five",
        "six",
        "seven",
        "eight",
        "nine",
        "ten",
        "zero",
    }

    # Apply formatting to both sheets containing the Gloss column
    for sheet_name in ["Details", "Gloss Source Stats"]:
        ws = wb[sheet_name]

        # Locate the Gloss column index dynamically
        gloss_col_idx = None
        for cell in ws[1]:
            if cell.value and str(cell.value).lower() == "gloss":
                gloss_col_idx = cell.column
                break

        if gloss_col_idx is None:
            continue

        # Iterate through data rows (skipping header)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=gloss_col_idx)
            val = str(cell.value).strip() if cell.value is not None else ""

            # Condition A: Check for digits or text-based number tokens
            tokens = set(re.split(r"[^a-zA-Z0-9]", val.lower()))
            has_digit = any(char.isdigit() for char in val)
            has_num_word = not tokens.isdisjoint(num_words)

            # Condition B: Check for a single English letter
            is_single_letter = len(val) == 1 and val.isalpha()

            if has_digit or has_num_word:
                cell.fill = light_blue
            elif is_single_letter:
                cell.fill = light_red

    wb.save(EXCEL_OUTPUT_PATH)
    print(f"Excel file generated successfully: {EXCEL_OUTPUT_PATH}")


if __name__ == "__main__":
    main()